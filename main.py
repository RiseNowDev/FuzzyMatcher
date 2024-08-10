import pandas as pd
from rapidfuzz import process, fuzz
from multiprocessing import Pool, cpu_count
from tqdm import tqdm
import time
import logging
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')

def normalize_name(name: str) -> str:
    return str(name).strip().lower()

def find_matches_and_scores(args: tuple) -> tuple:
    chunk, all_normalized_names, percentage_score, supplier_name, i = args
    matches = []
    for idx, row in chunk.iterrows():
        normalized_name = row['Normalized_Name']
        results = process.extract(
            normalized_name,
            all_normalized_names,
            scorer=fuzz.token_set_ratio,
            score_cutoff=percentage_score,
            limit=10
        )
        matches.append([(all_normalized_names[match[2]], match[2], match[1]) for match in results])
    return matches

def group_matches(data, matches):
    grouped_matches = []
    for idx, (row, match_list) in enumerate(zip(data.itertuples(), matches)):
        for match, match_idx, score in match_list:
            grouped_matches.append({
                'Original_Name': row.Normalized_Name,
                'Original_Index': getattr(row, 'i', idx),
                'Match_Name': match,
                'Match_Index': match_idx,
                'Score': score,
                'Suspicious': score < 95 or len(match) < 5  # Flag suspicious matches
            })
    return pd.DataFrame(grouped_matches)

def process_suppliers(input_file, output_file, num_cores=None, percentage_score=89,
                      supplier_name='supplier_name', i='i'):
    data = pd.read_csv(input_file)
    total_rows = len(data)
    num_cores = num_cores or cpu_count()

    logging.info(f"Total rows: {total_rows}")
    logging.info(f"Using {num_cores} cores")

    if 'Normalized_Name' not in data.columns:
        data['Normalized_Name'] = data[supplier_name].apply(normalize_name)

    all_normalized_names = data['Normalized_Name'].tolist()

    chunk_size = total_rows // num_cores
    data_chunks = [data[i:i + chunk_size] for i in range(0, total_rows, chunk_size)]

    args = [(chunk, all_normalized_names, percentage_score, supplier_name, i) for chunk in data_chunks]

    with Pool(processes=num_cores) as pool:
        results = list(tqdm(pool.imap(find_matches_and_scores, args), total=len(args), desc="Processing suppliers"))

    matches = [match for chunk_result in results for match in chunk_result]
    data['Potential_Matches'] = [', '.join([f"{match}|{idx}|{idx}|{score}" for match, idx, score in match_list]) for match_list in matches]
    
    # Create the grouped matches DataFrame
    grouped_matches_df = group_matches(data, matches)

    # Save both sheets to the Excel file
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        data.to_excel(writer, sheet_name='Original_Data', index=False)
        grouped_matches_df.to_excel(writer, sheet_name='Grouped_Matches', index=False)

    # Highlight suspicious matches
    highlight_suspicious_matches(output_file)

    logging.info(f"Results saved to {output_file}")

def highlight_suspicious_matches(file_path):
    workbook = load_workbook(file_path)
    sheet = workbook['Grouped_Matches']
    
    red_fill = PatternFill(start_color='FFFF0000', end_color='FFFF0000', fill_type='solid')
    
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=6, max_col=6):
        for cell in row:
            if cell.value == True:
                for col in range(1, 7):
                    sheet.cell(row=cell.row, column=col).fill = red_fill

    workbook.save(file_path)

if __name__ == "__main__":
    input_file = "raw/csusupps.csv"
    output_file = "output/csu_snorm.xlsx"

    percentage_score = float(input("Enter the percentage score to use for matching (default 89): ") or 89)
    name_field = input("Enter the name field to use (default supplier_name): ") or "supplier_name"
    i = input("Enter the name of the column which is your index. (default i): ") or "i"
    num_cores_input = input("Enter the number of cores to use (default max): ")
    num_cores = int(num_cores_input) if num_cores_input.strip() else None

    start_time = time.time()
    process_suppliers(input_file, output_file, num_cores, percentage_score, name_field, i)
    end_time = time.time()
    
    elapsed_time = end_time - start_time
    logging.info(f"Total processing time: {elapsed_time:.2f} seconds")